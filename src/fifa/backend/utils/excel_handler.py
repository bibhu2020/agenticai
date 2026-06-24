"""Excel read/write helpers for FIFA data (3 sheets: scores, schedule, news)."""
from __future__ import annotations
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_HEADER_FILL = PatternFill("solid", fgColor="1A3A5C")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_ALT_FILL = PatternFill("solid", fgColor="F0F4F8")
_CENTER = Alignment(horizontal="center", vertical="center")
_WRAP = Alignment(wrap_text=True, vertical="top")

SCORES_COLS = ["match_id", "date", "time_cdt", "home_team", "home_logo", "home_score",
               "away_score", "away_team", "away_logo", "status", "venue", "city", "round"]
SCHEDULE_COLS = ["match_id", "date", "time_cdt", "home_team", "home_logo", "away_team",
                 "away_logo", "venue", "city", "round", "status"]
NEWS_COLS = ["title", "summary", "thumbnail_url", "article_url", "published_date", "source"]


def _style_sheet(ws, col_widths: dict[str, int]) -> None:
    for col_idx, cell in enumerate(ws[1], start=1):
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        col_letter = get_column_letter(col_idx)
        header = cell.value or ""
        ws.column_dimensions[col_letter].width = col_widths.get(header, 15)
    ws.row_dimensions[1].height = 22


def create_excel(scores: list[dict], schedule: list[dict], news: list[dict]) -> bytes:
    wb = Workbook()

    ws_scores = wb.active
    ws_scores.title = "scores"
    ws_scores.append(SCORES_COLS)
    for m in scores:
        ws_scores.append([m.get(c, "") for c in SCORES_COLS])
    _style_sheet(ws_scores, {"match_id": 12, "date": 12, "time_cdt": 22,
                              "home_team": 20, "away_team": 20, "status": 14,
                              "venue": 25, "city": 18, "round": 20})

    ws_sched = wb.create_sheet("schedule")
    ws_sched.append(SCHEDULE_COLS)
    for m in schedule:
        ws_sched.append([m.get(c, "") for c in SCHEDULE_COLS])
    _style_sheet(ws_sched, {"match_id": 12, "date": 12, "time_cdt": 22,
                             "home_team": 20, "away_team": 20, "venue": 25,
                             "city": 18, "round": 20})

    ws_news = wb.create_sheet("news")
    ws_news.append(NEWS_COLS)
    for a in news:
        ws_news.append([a.get(c, "") for c in NEWS_COLS])
    _style_sheet(ws_news, {"title": 45, "summary": 60, "thumbnail_url": 40,
                            "article_url": 40, "published_date": 16, "source": 16})

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def read_excel(data: bytes) -> dict[str, list[dict]]:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)

    def _sheet_to_list(name: str) -> list[dict]:
        if name not in wb.sheetnames:
            return []
        ws = wb[name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h) if h is not None else "" for h in rows[0]]
        result = []
        for row in rows[1:]:
            if any(v is not None for v in row):
                result.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
        return result

    return {
        "scores": _sheet_to_list("scores"),
        "schedule": _sheet_to_list("schedule"),
        "news": _sheet_to_list("news"),
    }
